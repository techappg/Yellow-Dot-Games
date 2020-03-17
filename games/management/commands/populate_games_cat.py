from django.core.management import BaseCommand
from games.models import Category, Game
from openpyxl import load_workbook


class Command(BaseCommand):
    def handle(self, *args, **options):
        wb2 = load_workbook('games/All games.xlsx')
        i = 0
        for row in wb2['New-HTML'].iter_rows(min_row=1, values_only=True):
            if i == 0:
                i += 1
                continue
            cat = Category.objects.get_or_create(name=row[1], description=row[1], slug=row[1])
            Game.objects.get_or_create(name=row[2], category=cat[0], game_url=row[4], image_url=row[5])
            i += 1
